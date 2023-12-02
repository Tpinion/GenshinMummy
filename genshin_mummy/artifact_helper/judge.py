import logging
import os
from enum import Enum, unique
from logging import Logger
from typing import List, Optional, Sequence

import attrs
import openpyxl
from genshin_mummy.artifact_helper.exception import (
    BadArtifactType,
    BadConclusion,
    BadEntryOperator,
    BadEntryType,
    BadExcelHeader,
    BadNumber,
    BadNumberOperator,
    BadSeparator,
    LossActiveSheet,
    RedundantMainEntryType,
    RowValueError,
    SheetNotFound,
)
from genshin_mummy.artifact_helper.type import (
    Artifact,
    ArtifactType,
    EntryType,
)


@unique
class Conclusion(Enum):
    LOCK = '锁定'
    UNLOCK = '不锁'
    UNKNOWN = '没有决断'


@unique
class NumActor(Enum):
    GT: str = ">"
    LT: str = "<"
    EQ: str = "="


@unique
class EntryActor(Enum):
    POSITIVE: str = 'positive'
    NEGATIVE: str = 'negative'


@attrs.define
class NumCondition:
    actor: NumActor = attrs.field(default=None)
    num: int = attrs.field(default=None)

    def judge(self, value: int):
        if self.actor == NumActor.GT:
            return value > self.num
        elif self.actor == NumActor.LT:
            return value < self.num
        elif self.actor == NumActor.EQ:
            return value == self.num
        else:
            raise BadNumberOperator

    def __str__(self):
        return f"{self.actor.value}{self.num}"


@attrs.define
class EntryCondition:
    entry_type: EntryType = attrs.field()
    actor: EntryActor = attrs.field()

    def judge(self, candidate: EntryType):
        if self.actor == EntryActor.POSITIVE:
            return candidate == self.entry_type
        elif self.actor == EntryActor.NEGATIVE:
            return candidate != self.entry_type
        else:
            raise BadEntryOperator

    def all_not_match(self, entries: Sequence[EntryType]):
        for entry in entries:
            if self.judge(entry):
                return False
        return True

    def any_match(self, entries: Sequence[EntryType]):
        for entry in entries:
            if self.judge(entry):
                return True
        return False

    def __str__(self) -> str:
        if self.actor == EntryActor.POSITIVE:
            return f'有{self.entry_type.value}'
        elif self.actor == EntryActor.NEGATIVE:
            return f'没有{self.entry_type.value}'


@attrs.define
class MainEntryCondition:
    entry_condition: EntryCondition = attrs.field(factory=list)

    def judge(self, artifact: Artifact):
        entry = list(artifact.entry.keys())[0]
        return self.entry_condition.judge(entry)

    def __str__(self):
        if self.entry_condition.actor == EntryActor.POSITIVE:
            return f'是{self.entry_condition.entry_type.value}'
        else:
            return f'不是{self.entry_condition.entry_type.value}'


@attrs.define
class SubEntryCondition:
    entry_conditions: Sequence[EntryCondition] = attrs.field(factory=list)

    def judge(self, artifact: Artifact):
        subentries = list(artifact.subentries.keys())
        for entry_condition in self.entry_conditions:
            if (entry_condition.actor == EntryActor.POSITIVE
                    and entry_condition.all_not_match(subentries)):
                return False
            elif (entry_condition.actor == EntryActor.NEGATIVE
                  and entry_condition.any_match(subentries)):
                return False
        return True

    def __str__(self):
        return '、'.join([str(mem) for mem in self.entry_conditions])


@attrs.define
class Strategy:
    artifact_types: Optional[Sequence[ArtifactType]] = attrs.field()
    level_condition: Optional[NumCondition] = attrs.field()
    star_condition: Optional[NumCondition] = attrs.field()
    main_entry_condition: Optional[MainEntryCondition] = attrs.field()
    subentries_condition: Optional[SubEntryCondition] = attrs.field()
    conclusion: Conclusion = attrs.field()

    conclusion: bool = attrs.field()

    def apply(self, artifact: Artifact):
        if self.artifact_types and artifact.type not in self.artifact_types:
            return False
        if self.level_condition and not self.level_condition.judge(
                artifact.level):
            return False
        if self.star_condition and not self.star_condition.judge(
                artifact.stars):
            return False
        if self.main_entry_condition and not self.main_entry_condition.judge(
                artifact):
            return False
        if self.subentries_condition and not self.subentries_condition.judge(
                artifact):
            return False
        return True

    def __str__(self):
        message = f'[{self.conclusion.value}]  '
        if self.artifact_types:
            message += '、'.join([mem.value for mem in self.artifact_types])
        else:
            message += "任意圣遗物类型"

        if self.level_condition:
            message += f"，等级{self.level_condition}"
        else:
            message += "，任意等级"

        if self.star_condition:
            message += f"，星级{self.star_condition}"
        else:
            message += "，任意星级"

        if self.main_entry_condition:
            message += f"，主词条{self.main_entry_condition}"
        else:
            message += "，任意主词条"

        if self.subentries_condition:
            message += f"，副词条{str(self.subentries_condition)}"
        else:
            message += "，任意副词条"

        return message


class StrategyFactory:
    SEP = ','
    DEFAULT_SHEET_NAME = 'Active'
    HEADERS = ["圣遗物类型条件", "等级条件", "星级条件", "主词条条件", "副词条条件", "期望结果"]

    def __init__(self, config_excel_fp: str, logger: Logger) -> None:
        self.config_excel_fp = config_excel_fp
        self.logger = logger
        self.acl = []
        self.artifact_type_value_to_member = {
            member.value: member
            for member in ArtifactType
        }
        self.entry_type_value_to_member = {
            member.value: member
            for member in EntryType
        }

    def check_header(self, header_row):
        for std_header, cell in zip(self.HEADERS, header_row):
            if std_header != cell.value:
                raise BadExcelHeader

    def check_sep(self, value: str):

        def is_chinese(char: str):
            return '\u4e00' <= char <= '\u9fff'

        for char in value:
            if not is_chinese(char) and char != self.SEP:
                raise BadSeparator

    def get_entry_type(self, value: str):
        try:
            return self.entry_type_value_to_member[value]
        except KeyError:
            raise BadEntryType(value)

    def get_artifact_type(self, value: str):
        try:
            return self.artifact_type_value_to_member[value]
        except KeyError:
            raise BadArtifactType(
                artifact_type=value,
                tag='圣遗物类型条件',
            )

    def format(self, value):
        if value:
            return value.replace(" ", "").replace("，", self.SEP)
        return ""

    def load_acl(self):
        acl = []
        wb = openpyxl.load_workbook(self.config_excel_fp)

        if not wb.sheetnames:
            raise SheetNotFound
        elif len(wb.sheetnames) == 1:
            active_sheet = wb[wb.sheetnames[0]]
        else:
            if self.DEFAULT_SHEET_NAME not in wb.sheetnames:
                raise LossActiveSheet
            active_sheet = wb[self.DEFAULT_SHEET_NAME]

        active_sheet = wb['Active']
        for index, row in enumerate(active_sheet.rows):
            if index == 0:
                self.check_header(row)
                continue

            col_values = [self.format(cell.value) for cell in row]
            acl_item = dict(zip(self.HEADERS, col_values))
            acl.append(acl_item)
        self.acl = acl

    def translate_artifact_types(self, value: str):
        if value == "":
            return None
        artifact_type_names = value.split(self.SEP)
        artifact_types = []
        for name in artifact_type_names:
            artifact_type = self.get_artifact_type(name)
            artifact_types.append(artifact_type)
        return artifact_types

    def translate_num_condition(self, value: str):
        if value == "":
            return None
        sep_index = None
        for idx, char in enumerate(value):
            if char.isdigit():
                sep_index = idx
                break
        actor_str = value[:sep_index]

        try:
            level = int(value[sep_index:])
        except ValueError:
            raise BadNumber

        if actor_str == NumActor.GT.value:
            return NumCondition(actor=NumActor.GT, num=int(level))
        elif actor_str == NumActor.LT.value:
            return NumCondition(actor=NumActor.LT, num=int(level))
        else:
            raise BadNumberOperator

    def translate_level_condition(self, value: str):
        return self.translate_num_condition(value)

    def translate_star_condition(self, value: str):
        return self.translate_num_condition(value)

    def translate_entry_condition(
        self,
        value: str,
        pos_word: str,
        neg_word: str,
    ):
        self.check_sep(value)
        items = value.split(self.SEP)
        entry_connditions = []
        for item in items:
            if item.startswith(neg_word):
                entry_value = item[len(neg_word):]
                entry_connditions.append(
                    EntryCondition(
                        entry_type=self.get_entry_type(entry_value),
                        actor=EntryActor.NEGATIVE,
                    ))
            elif item.startswith(pos_word):
                entry_value = item[1:]
                entry_connditions.append(
                    EntryCondition(
                        entry_type=self.get_entry_type(entry_value),
                        actor=EntryActor.POSITIVE,
                    ))
            else:
                raise BadEntryOperator(pos_word, neg_word)
        return entry_connditions

    def translate_main_entry_condition(self, value: str):
        POSITIVE_WORD = '是'
        NEGATIVE_WORD = '不是'
        if value == "":
            return None
        try:
            entry_connditions = self.translate_entry_condition(
                value,
                POSITIVE_WORD,
                NEGATIVE_WORD,
            )
            if len(entry_connditions) > 1:
                raise RedundantMainEntryType
            entry_conndition = entry_connditions[0]
        except (BadEntryType, BadEntryOperator) as error:
            error.tag = '主词条条件'
            raise error

        return MainEntryCondition(entry_condition=entry_conndition)

    def translate_subentries_condition(self, value: str):
        if value == "":
            return None
        POSITIVE_WORD = '有'
        NEGATIVE_WORD = '没有'
        try:
            entry_connditions = self.translate_entry_condition(
                value,
                POSITIVE_WORD,
                NEGATIVE_WORD,
            )
        except (BadEntryType, BadEntryOperator) as error:
            error.tag = '副词条条件'
            raise error
        return SubEntryCondition(entry_conditions=entry_connditions)

    def translate_conclusion(self, value: str):
        if value == "锁":
            return Conclusion.LOCK
        elif value == "不锁":
            return Conclusion.UNLOCK
        else:
            raise BadConclusion(value)

    def translate_acl(self):
        strategies: List[Strategy] = []
        for index, item in enumerate(self.acl):
            try:
                artifact_types = self.translate_artifact_types(
                    item[self.HEADERS[0]])
                level_condition = self.translate_level_condition(
                    item[self.HEADERS[1]])
                star_condition = self.translate_star_condition(
                    item[self.HEADERS[2]])
                main_entry_condition = self.translate_main_entry_condition(
                    item[self.HEADERS[3]])
                subentries_condition = self.translate_subentries_condition(
                    item[self.HEADERS[4]])

                conclusion = self.translate_conclusion(item[self.HEADERS[5]])
            except RowValueError as error:
                error.line = index + 2
                raise error

            strategies.append(
                Strategy(
                    artifact_types=artifact_types,
                    level_condition=level_condition,
                    star_condition=star_condition,
                    main_entry_condition=main_entry_condition,
                    subentries_condition=subentries_condition,
                    conclusion=conclusion,
                ))
        return strategies


class ArtifactJudge:

    def __init__(
        self,
        config_fp: Optional[str] = None,
        logger: Optional[Logger] = None,
    ):
        if logger is None:
            logging.basicConfig(
                format='%(levelname)s - %(message)s',
                level=logging.INFO,
            )
            logger = logging.getLogger(__name__)
            logger.setLevel(logging.INFO)

        self.logger = logger
        self.strategys: Sequence[Strategy] = []

        if config_fp and os.path.exists(config_fp):
            factory = StrategyFactory(config_fp, logger)
            factory.load_acl()
            self.strategys = factory.translate_acl()

    def judge(self, artifact: Artifact):
        self.logger.info("开启圣遗物加解锁判断...")
        if self.strategys:
            conclusion = self.user_strategy(artifact)
        else:
            conclusion = self.default_strategy(artifact)
        return conclusion

    def user_strategy(self, artifact):
        self.logger.info("开始执行用户策略...")
        for index, strategy in enumerate(self.strategys):
            if strategy.apply(artifact):
                self.logger.info((f"命中第{index+2}条策略： {str(strategy)}"))
                return strategy.conclusion
        self.logger.info("未命中任何策略")
        return Conclusion.UNKNOWN

    def default_strategy(self, artifact):
        self.logger.info("开始执行默认策略...")
        # 算法参考
        # https://www.bilibili.com/video/BV1sZ4y1e7h8
        # https://www.bilibili.com/video/BV1mB4y177a6

        # 等级大于0=>锁
        if artifact.level > 0:
            return Conclusion.LOCK

        # 非五星=>不锁
        if artifact.stars < 5:
            return Conclusion.UNLOCK

        # 沙、杯、帽主词条为类别独有词条=>锁
        if (artifact.type not in {
                ArtifactType.FLOWER_OF_LIFE, ArtifactType.PLUME_OF_DEATH
        } and list(artifact.entry.keys())[0] not in {
                EntryType.HP_PERCENTAGE, EntryType.ATK_PERCENTAGE,
                EntryType.DEF_PERCENTAGE
        }):
            return Conclusion.LOCK

        # 双暴词条=>锁
        subentry_types = set(artifact.subentries.keys())
        if {EntryType.CRIT_DMG, EntryType.CRIT_RATE}.issubset(subentry_types):
            return Conclusion.LOCK

        # 初始四词条 且不要存在所有小攻防命都有=>锁
        if len(subentry_types) == 4 and not {
                EntryType.HP, EntryType.ATK, EntryType.DEF
        }.issubset(subentry_types):
            return Conclusion.LOCK

        # 小攻击、小防御、小生命大于等于两个=>不锁
        if len({EntryType.HP, EntryType.ATK, EntryType.DEF}
               & subentry_types) >= 2:
            return Conclusion.UNLOCK

        return Conclusion.LOCK
