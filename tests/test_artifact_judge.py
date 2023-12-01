from genshin_mummy.artifact_helper.judge import ArtifactJudge, Conclusion
from genshin_mummy.artifact_helper.type import (
    Artifact,
    EntryType,
    ArtifactType,
)

from genshin_mummy.artifact_helper.exception import (
    BadExcelHeader,
    BadArtifactType,
    BadEntryType,
    BadSeparator,
)

import openpyxl
import tempfile


def test_bad_sheet_header():
    excel_rows = [
        ["圣遗物类型", "等级条件", "星级条件", "主词条条件", "副词条条件", "期望结果"],
        [None, None, None, None, "有攻击力，有防御力", "不锁"],
        [None, None, None, None, None, "锁"],
    ]

    fd, fp = tempfile.mkstemp(suffix='.xlsx')

    with open(fp, 'w') as fout:
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet("Active")
        for row in excel_rows:
            sheet.append(row)
        workbook.save(fout.name)

    try:
        ArtifactJudge(fout.name)
        raise Exception("Should not reach here")
    except BadExcelHeader:
        pass


def test_bad_artifact_type():
    excel_rows = [
        ["圣遗物类型条件", "等级条件", "星级条件", "主词条条件", "副词条条件", "期望结果"],
        ["花", None, None, None, "有攻击力、有防御力", "不锁"],
        [None, None, None, None, None, "锁"],
    ]

    fd, fp = tempfile.mkstemp(suffix='.xlsx')

    with open(fp, 'w') as fout:
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet("Active")
        for row in excel_rows:
            sheet.append(row)
        workbook.save(fout.name)

    try:
        ArtifactJudge(fout.name)
        raise Exception("Should not reach here")
    except BadArtifactType:
        pass


def test_bad_separator():
    excel_rows = [
        ["圣遗物类型条件", "等级条件", "星级条件", "主词条条件", "副词条条件", "期望结果"],
        [None, None, None, None, "有攻击力、有防御力", "不锁"],
        [None, None, None, None, None, "锁"],
    ]

    fd, fp = tempfile.mkstemp(suffix='.xlsx')

    with open(fp, 'w') as fout:
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet("Active")
        for row in excel_rows:
            sheet.append(row)
        workbook.save(fout.name)

    try:
        ArtifactJudge(fout.name)
        raise Exception("Should not reach here")
    except BadSeparator:
        pass


def test_bad_entry_type_1():
    excel_rows = [
        ["圣遗物类型条件", "等级条件", "星级条件", "主词条条件", "副词条条件", "期望结果"],
        [None, None, None, None, "有攻击力，有防御", "不锁"],
        [None, None, None, None, None, "锁"],
    ]

    fd, fp = tempfile.mkstemp(suffix='.xlsx')

    with open(fp, 'w') as fout:
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet("Active")
        for row in excel_rows:
            sheet.append(row)
        workbook.save(fout.name)

    try:
        ArtifactJudge(fout.name)
        raise Exception("Should not reach here")
    except BadEntryType:
        pass


def test_bad_entry_type_2():
    excel_rows = [
        ["圣遗物类型条件", "等级条件", "星级条件", "主词条条件", "副词条条件", "期望结果"],
        [None, None, None, "是防御", "有攻击力，有防御力", "不锁"],
        [None, None, None, None, None, "锁"],
    ]

    fd, fp = tempfile.mkstemp(suffix='.xlsx')

    with open(fp, 'w') as fout:
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet("Active")
        for row in excel_rows:
            sheet.append(row)
        workbook.save(fout.name)

    try:
        ArtifactJudge(fout.name)
        raise Exception("Should not reach here")
    except BadEntryType:
        pass


def test_shit_1():
    excel_rows = [
        ["圣遗物类型条件", "等级条件", "星级条件", "主词条条件", "副词条条件", "期望结果"],
        [None, None, None, None, "有攻击力，有防御力", "不锁"],
        [None, None, None, None, None, "锁"],
    ]

    fd, fp = tempfile.mkstemp(suffix='.xlsx')

    with open(fp, 'w') as fout:
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet("Active")
        for row in excel_rows:
            sheet.append(row)
        workbook.save(fout.name)

    artifact = Artifact(
        name='赌徒',
        type=ArtifactType.FLOWER_OF_LIFE,
        entry={EntryType.HP_PERCENTAGE: "22"},
        stars=5,
        level=20,
        subentries={
            EntryType.HP: "22",
            EntryType.ATK: "22",
            EntryType.DEF: "22",
        },
    )

    jduge = ArtifactJudge(fout.name)
    conclusion = jduge.judge(artifact)
    assert conclusion == Conclusion.UNLOCK


def test_shit_2():
    excel_rows = [
        ["圣遗物类型条件", "等级条件", "星级条件", "主词条条件", "副词条条件", "期望结果"],
        [None, None, None, None, "有防御力，有防御力百分比", "不锁"],
        [None, None, None, None, "有防御力，有生命值", "不锁"],
        [None, None, None, None, None, "锁"],
    ]

    fd, fp = tempfile.mkstemp(suffix='.xlsx')

    with open(fp, 'w') as fout:
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet("Active")
        for row in excel_rows:
            sheet.append(row)
        workbook.save(fout.name)

    artifacts = [
        (Artifact(
            name='角斗士的留恋',
            type=ArtifactType.FLOWER_OF_LIFE,
            entry={EntryType.HP_PERCENTAGE: "22"},
            stars=5,
            level=20,
            subentries={
                EntryType.HP: "22",
                EntryType.ATK: "22",
                EntryType.DEF: "22",
            },
        ), Conclusion.UNLOCK),
        (Artifact(
            name='角斗士的留恋',
            type=ArtifactType.FLOWER_OF_LIFE,
            entry={EntryType.HP_PERCENTAGE: "22"},
            stars=5,
            level=20,
            subentries={
                EntryType.HP_PERCENTAGE: "22",
                EntryType.ATK: "22",
                EntryType.DEF: "22",
            },
        ), Conclusion.LOCK),
    ]

    jduge = ArtifactJudge(fout.name)
    for artifact, conclusion in artifacts:
        assert jduge.judge(artifact) == conclusion
